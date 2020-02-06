import numpy as np
import cv2
import pywt
import glob
from PIL import Image
import copy
import openpyxl
import math
from matplotlib import pyplot as plt


def mse(I1, I2):
    err = np.square(np.subtract(np.double(I1), np.double(I2))).mean()
    return err


def psnr(img1, img2):
    mseval = mse(img1, img2)
    if mse == 0:
        return 100
    PIXEL_MAX = 255.0
    return 20 * math.log10(PIXEL_MAX / math.sqrt(mseval))


def halt():
    k = 0
    while k != 48:
        k = cv2.waitKey(100)
    cv2.destroyAllWindows()


def Save_Image(Location, Title, FName2, image_array):
    fn = Title + FName2
    img = Image.fromarray(image_array)
    Floc = Location + fn
    img.save(Floc)  # save image at location
    print('Image Saved at: ' + Floc)
    return


def DFT(coverImage, watermarkImage):
    print("--- DFT Watermarking ---")
    H, W = np.shape(coverImage)  # find size of cover image
    coverImage = coverImage / 255
    coverImageFFT = np.fft.fftshift(np.fft.fft2(coverImage))
    Hf, Wf = np.shape(coverImageFFT)

    Hw, Ww = np.shape(watermarkImage)
    watermarkImage2 = cv2.resize(watermarkImage, (W, H))
    Hw2, Ww2 = np.shape(coverImage)
    watermarkImage2 = watermarkImage2 / 255
    watermarkImageFFT = np.fft.fftshift(np.fft.fft2(watermarkImage2))
    Hf, Wf = np.shape(watermarkImageFFT)

    alpha = 0.05
    wattermarkedFFT = coverImageFFT + (alpha * watermarkImageFFT)

    watermarkedImage = np.fft.ifft2(np.fft.ifftshift(wattermarkedFFT))

    ## calculate MSE
    watermarkedImage = np.uint8(255 * watermarkedImage)
    coverImage2 = copy.deepcopy(coverImage)
    coverImage2 = np.uint8(255 * coverImage)

    MSE = mse(coverImage2, watermarkedImage)
    print('MSE (coverimage, watermarked)=' + str(MSE))
    PSNR = psnr(coverImage2, watermarkedImage)
    print('PSNR (coverimage, watermarked)=' + str(PSNR))
    ## UPDATE EXCEL DATA
    Technique = 'DFT'
    XLS_path = "watermarking results.xlsx"
    XLS_Obj = openpyxl.load_workbook(XLS_path)
    XLS_Sheet_Obj = XLS_Obj.active
    CountROW = XLS_Sheet_Obj.max_row
    CountCol = 1
    print('ROW Location=' + str(CountROW))
    CELL1 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=1);
    CELL1.value = FName
    CELL2 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=2);
    CELL2.value = 'DFT'
    CELL3 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=3);
    CELL3.value = str(MSE)
    CELL4 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=4);
    CELL4.value = str(PSNR)
    XLS_Obj.save("watermarking results.xlsx")


    # save watermarked image
    fn = 'DFT_Watermarked_' + FName  # set file name
    # cv2.imshow('DFT_Watermarked_' + FName, watermarkedImage)

    img = Image.fromarray(watermarkedImage)
    Floc = './result/DFT/' + fn  # set file path+file name
    img.save(Floc)  # save image at location
    print('Watermarked Image saved at:' + Floc)

    return


def DWT_Encoding(coverImage, watermarkImage, alpha):
    coeffC = pywt.dwt2(coverImage, 'haar')
    cA, (cH, cV, cD) = coeffC

    # Embedding
    coeffW = (cA + alpha * watermarkImage, (cH, cV, cD))
    # Apply inverse DWT
    watermarkedImage = pywt.idwt2(coeffW, 'haar')

    watermarkedImage = np.uint8(255 * watermarkedImage)
    coverImage = np.uint8(255 * coverImage)
    print("Embedding Watermark done..")
    # cv2.imshow('DWT Watermarked Image: '+FName, watermarkedImage)

    return (watermarkedImage, cA)


def DWT_Decoding(twatermarkedImage, alpha, cA):
    # Extraction
    twatermarkedImage = twatermarkedImage / 255
    coeffWM = pywt.dwt2(twatermarkedImage, 'haar')
    hA, (hH, hV, hD) = coeffWM

    textracted = (hA - cA) / alpha
    textracted = np.uint8(255 * textracted)

    return (textracted)


def DWT(coverImage, watermarkImage):
    print("--- DWT Watermarking ---")

    coverImage = cv2.resize(coverImage, (300, 300))
    # cv2.imshow('Cover Image: '+FName,coverImage)
    watermarkImage = cv2.resize(watermarkImage, (150, 150))
    # cv2.imshow('Watermark Image',watermarkImage)

    # DWT on cover image
    coverImage = np.float32(coverImage)
    coverImage /= 255

    watermarkImage = np.float32(watermarkImage)
    watermarkImage /= 255
    alpha = 0.1
    ## ENCODING DWT
    watermarkedImage, cA = DWT_Encoding(coverImage, watermarkImage, alpha)

    MSE = mse(coverImage, watermarkedImage);
    print('MSE (coverimage, watermarked cover image)=' + str(MSE))
    PSNR = psnr(coverImage, watermarkedImage);
    print('PSNR (coverimage, watermarked cover image)=' + str(PSNR))

    ## UPDATE EXCEL DATA
    Technique = 'DWT'
    XLS_path = "watermarking results.xlsx"
    XLS_Obj = openpyxl.load_workbook(XLS_path)
    XLS_Sheet_Obj = XLS_Obj.active
    CountROW = XLS_Sheet_Obj.max_row
    CountCol = 1
    print('ROW Location=' + str(CountROW))
    CELL1 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=1);
    CELL1.value = FName
    CELL2 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=2);
    CELL2.value = 'DWT'
    CELL3 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=3);
    CELL3.value = str(MSE)
    CELL4 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=4);
    CELL4.value = str(PSNR)
    XLS_Obj.save("watermarking results.xlsx")
    # save watermarked image
    Save_Image('./result/DWT/', 'DWT_Watermarked_', FName, watermarkedImage)

    return watermarkedImage


def DCT(I, Imask):
    print("--- DCT Watermarking ---")
    [Hw, Ww] = np.shape(Imask)
    NWp = Hw * Ww
    print('Watermark Image Size=' + str(Hw) + 'x' + str(Ww) + '=' + str(NWp))

    # I = cv2.imread('D:\PROJECT\python\code\DCTTest\coverImage1.jpg', 0)
    # cv2.imshow('Cover Image: '+FName, I)
    [H, W] = np.shape(I)
    print('Cover Image=' + str(H) + 'x' + str(W))

    ## Lower DCT Coefficient locations of 8x8 block
    RC = [[0, 0], [1, 0], [2, 0], [3, 0], [4, 0], [5, 0], [6, 0],
          [0, 1], [1, 1], [2, 1], [3, 1], [4, 1], [5, 1],
          [0, 2], [1, 2], [2, 2], [3, 2], [4, 2],
          [0, 3], [1, 3], [2, 3], [3, 3],
          [0, 4], [1, 4], [2, 4],
          [0, 5], [1, 5],
          [0, 6]]
    NALoc = RC.__len__()
    print('Low Coef N=' + str(NALoc))

    ## Total pixels tobe processed
    Bh = np.uint8(H / 8)
    Bw = np.uint8(W / 8)
    TP = np.double(Bh) * np.double(Bw) * np.double(NALoc)
    print('Total Pixels can be processed=', str(TP))

    ## Reduce size of watermark image
    while (TP < NWp):  # if watermark image is too large, reduce its size
        Nsize = np.uint16(np.round(0.7 * np.double(np.shape(Imask))))
        # Imask=np.resize(Imask,Nsize)
        Imask = cv2.resize(Imask, (Nsize[0], Nsize[1]), interpolation=cv2.INTER_AREA)
        [Hw, Ww] = np.shape(Imask)
        NWp = Hw * Ww
        print('Watermark Pixels=' + str(Hw) + 'x' + str(Ww) + '=' + str(NWp))

    Imask = cv2.threshold(Imask, 127, 255, cv2.THRESH_BINARY)[1]
    ##cv2.imshow('Resized Watermark Image: ', Imask)
    Imask = Imask / 255
    # cv2.imshow('BW Resized Watermark Image: ', Imask)

    ## Reshape watermark image to an single array
    ImaskA = np.reshape(Imask, (NWp, 1))
    R = NWp % NALoc
    Zp = NALoc - R
    ZAp = np.zeros((Zp, 1))
    ImaskA = np.concatenate((ImaskA, ZAp))  # combine in row direction
    NWp = ImaskA.size

    ##process blocks
    Wc = 0
    Iw = copy.deepcopy(I)

    alpha = 2
    DctOrg = np.zeros((NWp, 1))
    DctVal = np.zeros((NALoc, 1))
    for i in range(0, H - 8, 8):
        for j in range(0, W - 8, 8):

            Iblock = I[i:i + 8, j: j + 8]  # crop image block 8x8
            DctBlock = cv2.dct(np.double(Iblock))  # convert to dct

            for k in range(0, NALoc):
                r = RC[k][0]
                c = RC[k][1]
                val = DctBlock[r, c]
                DctVal[k] = val  # find lower dc component

            Wc2 = Wc + NALoc  # end location
            # if(i==0):
            #       print('i=' + str(i) + ', ' + 'j=' + str(j)+', Wc='+str(Wc)+', Wc2='+str(Wc2))

            DctOrg[Wc: Wc2] = DctVal  # copy original data

            MaskVal = ImaskA[Wc:Wc2]  # copy watermark pixel
            DctVal2 = DctVal + alpha * MaskVal  # Embed watermark

            for k in range(0, NALoc):
                r = RC[k][0]
                c = RC[k][1]
                val = DctVal2[k]
                DctBlock[r, c] = val  # find lower dc component end

            Block2 = cv2.idct(DctBlock)  # apply inverse dct
            Iw[i: i + 8, j: j + 8] = Block2  # copy back to image
            if (Wc > (NWp - NALoc - 1)):
                break
            else:
                Wc = Wc + NALoc
        if (Wc > (NWp - NALoc)):
            break
    # Display watermarked image

    Iw = np.round(Iw)
    # cv2.imshow('Watermarked Image: '+FName, Iw)

    # MSE = np.square(np.subtract(np.double(I),np.double(Iw))).mean()
    MSE = mse(I, Iw)
    print('MSE (coverimage, watermarked cover image)=' + str(MSE))
    PSNR = psnr(I, Iw)
    print('PSNR (coverimage, watermarked cover image)=' + str(PSNR))

    ## UPDATE EXCEL DATA
    Technique = 'DCT'
    XLS_path = "watermarking results.xlsx"
    XLS_Obj = openpyxl.load_workbook(XLS_path)
    XLS_Sheet_Obj = XLS_Obj.active
    CountROW = XLS_Sheet_Obj.max_row
    CountCol = 1
    print('ROW Location=' + str(CountROW))
    CELL1 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=1);
    CELL1.value = FName
    CELL2 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=2);
    CELL2.value = 'DCT'
    CELL3 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=3);
    CELL3.value = str(MSE)
    CELL4 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=4);
    CELL4.value = str(PSNR)
    XLS_Obj.save("watermarking results.xlsx")

    # save watermarked image
    img = Image.fromarray(Iw)
    fn = 'DCT_Watermarked_' + FName  # set file name
    Floc = './result/DCT/' + fn  # set file path+file name
    img.save(Floc)  # save image at location
    print('Watermarked Image saved at: ' + Floc)

    return


def SVD(coverImage, watermarkImage):
    print("--- SVD Watermarking ---")
    # cv2.imshow('Cover Image: '+FName,coverImage)
    [H, W] = np.shape(coverImage)
    m = H
    n = W
    if n < m:
        m = n
    else:
        n = m

    coverImage2 = cv2.resize(coverImage, (n, m))  # resize to size of cover image
    coverImage2 = np.double(coverImage2)
    # plt.imshow(coverImage2,cmap='gray');plt.title('Resized cover image');plt.show()

    watermarkImage2 = copy.deepcopy(watermarkImage)
    watermarkImage2 = cv2.resize(watermarkImage2, (n, m))  # resize to size of cover image
    # plt.imshow(watermarkImage2,cmap='gray');plt.title('Resized watermark image');plt.show()

    watermarkImageBW = cv2.threshold(watermarkImage2, 127, 255, cv2.THRESH_BINARY)[1]
    watermarkImageBW = watermarkImageBW / 255  # convert to 0,1 range

    # SVD of cover image
    Uimg, Simg, Vimg = np.linalg.svd(coverImage2, full_matrices=1, compute_uv=1)
    Simg = np.diag(Simg)  # convert to diagonal matrix
    Simg_temp = copy.deepcopy(Simg)  # copy to new variable

    # Apply watermarking
    alpha = 10
    Simg = Simg + (alpha * watermarkImageBW)  # update simg

    # SVD of watermarked Simg
    U_SHL_w, S_SHL_w, V_SHL_w = np.linalg.svd(Simg, full_matrices=1, compute_uv=1)

    ## Applying inverse SVD
    watermarkedImage = np.uint8(np.dot((Uimg * S_SHL_w), Vimg))
    # plt.imshow(watermarkedImage,cmap='gray');plt.title('Watermarked processed image');plt.show()
    watermarkedImage = cv2.resize(watermarkedImage, (W, H))
    # plt.imshow(watermarkedImage,cmap='gray');plt.title('Watermarked resizeds image');plt.show()

    err = mse(coverImage, watermarkedImage)  # calculate mse
    print('MSE (cover, watermarked)=' + str(err))
    PSNR = psnr(coverImage, watermarkedImage)
    print('PSNR (cover, watermarked)=' + str(PSNR))

    ## UPDATE EXCEL DATA
    Technique = 'SVD'
    XLS_path = "watermarking results.xlsx"
    XLS_Obj = openpyxl.load_workbook(XLS_path)
    XLS_Sheet_Obj = XLS_Obj.active
    CountROW = XLS_Sheet_Obj.max_row
    CountCol = 1
    print('ROW Location=' + str(CountROW))
    CELL1 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=1);
    CELL1.value = FName
    CELL2 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=2);
    CELL2.value = 'SVD'
    CELL3 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=3);
    CELL3.value = str(err)
    CELL4 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=4);
    CELL4.value = str(PSNR)
    XLS_Obj.save("watermarking results.xlsx")
    # save watermarked image
    watermarkedImage = np.uint8(watermarkedImage)
    fn = 'SVD_Watermarked_' + FName  # set file name
    # cv2.imshow('SVD_Watermarked_' + FName, watermarkedImage)

    img = Image.fromarray(watermarkedImage)
    Floc = './result/SVD/' + fn  # set file path+file name
    img.save(Floc)  # save image at location
    print('Watermarked Image saved at: ' + Floc)

    return


# option 5
def DWT_SVD(coverImage, watermarkImage):
    print("--- DWT SVD Watermarking ---")
    [Hgt, Wdt] = np.shape(coverImage)
    m = Hgt
    n = Wdt
    if n < m:
        m = n
    else:
        n = m

    coverImage2 = cv2.resize(coverImage, (n, m))  # resize to size of cover image
    coverImage2 = np.double(coverImage2)
    # plt.imshow(coverImage2, cmap='gray');plt.title('Cover image resized');plt.show()

    # Apply DWT
    coeffC = pywt.dwt2(coverImage2, 'haar')
    A, (H, V, D) = coeffC
    [m, n] = np.shape(A)

    watermarkImage2 = copy.deepcopy(watermarkImage)
    watermarkImage2 = cv2.resize(watermarkImage2, (n, m))  # resize to size of cover image
    watermarkImageBW = cv2.threshold(watermarkImage2, 127, 255, cv2.THRESH_BINARY)[1]
    # plt.imshow(watermarkImageBW, cmap='gray');plt.title('resized watermark image');plt.show()
    watermarkImageBW = watermarkImageBW / 255  # convert to 0,1 range

    # Apply SVD on A Band
    Uimg, Simg, Vimg = np.linalg.svd(A, full_matrices=1, compute_uv=1)
    Simg = np.diag(Simg)  # convert to diagonal matrix
    Simg_temp = copy.deepcopy(Simg)  # copy to new variable

    # Apply watermarking
    alpha = 10
    Simg = Simg + (alpha * watermarkImageBW)  # update simg

    # SVD of watermarked Simg
    U_SHL_w, S_SHL_w, V_SHL_w = np.linalg.svd(Simg, full_matrices=1, compute_uv=1)

    ## Applying inverse SVD
    A2 = np.dot((Uimg * S_SHL_w), Vimg)
    # Perform inverse dwt and generate watermarked image
    coeffW = A2, (H, V, D)
    watermarkedImage = pywt.idwt2(coeffW, 'haar')
    watermarkedImage = np.round_(watermarkedImage)
    watermarkedImage = np.uint8(watermarkedImage)

    watermarkedImage = cv2.resize(watermarkedImage, (Wdt, Hgt))
    # plt.imshow(watermarkedImage,cmap='gray');plt.title('Resized Watermarked image');plt.show()

    # CALCULATE MSE
    err = mse(coverImage, watermarkedImage)
    print('MSE (cover image, watermarked image)=', err)
    PSNR = psnr(coverImage, watermarkedImage)
    print('PSNR (cover image, watermarked image)=' + str(PSNR))

    ## UPDATE EXCEL DATA
    Technique = 'DWT_SVD'
    XLS_path = "watermarking results.xlsx"
    XLS_Obj = openpyxl.load_workbook(XLS_path)
    XLS_Sheet_Obj = XLS_Obj.active
    CountROW = XLS_Sheet_Obj.max_row
    CountCol = 1
    print('ROW Location=' + str(CountROW))
    CELL1 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=1);
    CELL1.value = FName
    CELL2 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=2);
    CELL2.value = 'DWT_SVD'
    CELL3 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=3);
    CELL3.value = str(err)
    CELL4 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=4);
    CELL4.value = str(PSNR)
    XLS_Obj.save("watermarking results.xlsx")

    # save watermarked image
    img = Image.fromarray(watermarkedImage)
    fn = 'DWTSVD_Watermarked_' + FName  # set file name
    Floc = './result/DWT_SVD/' + fn  # set file path+file name
    img.save(Floc)  # save image at location
    print('Watermarked Image saved at:' + Floc)
    return watermarkedImage


# Option 6
def DWT_DCT_SVD(coverImage, watermarkImage):
    print('--- DCT Watermarking ---')
    coverImage = cv2.imread(st, 0)  # read cover image
    coverImage = cv2.resize(coverImage, (512, 512))
    # cv2.imshow('Cover Image : '+FName,coverImage)

    # Read watermark image
    watermarkImage = cv2.imread('watermarkImage4.PNG', 0)
    watermarkImage = cv2.resize(watermarkImage, (32, 32))
    # cv2.namedWindow("Watermark Image", cv2.WINDOW_NORMAL)
    # cv2.imshow('Watermark Image',watermarkImage)

    print("Read watermark image")

    # Apply DWT
    coeffC = pywt.dwt2(coverImage, 'haar')
    A, (H, V, D) = coeffC
    print("DWT of coverimage done");
    [Ha, Wa] = np.shape(A)

    # Generate DCT Block
    Bh = np.uint8(Ha / 8)  # total block can be processed in horizontal direction
    Bw = np.uint8(Wa / 8)  # total blocks can be processed in vertical direction

    DMat = np.zeros((Bh, Bw), np.float32)
    dblock = 7
    ih = 0
    iw = 0
    for i in range(0, Ha - 1, 8):
        iw = 0
        for j in range(0, Wa - 1, 8):
            i2 = i + 8;
            j2 = j + 8;
            Block = A[i:i2, j:j2]  # crop the 8x8 image block
            DctBlock = cv2.dct(Block)  # convert to dct
            # DMat[ih, iw]=DctBlock[dblock, dblock]
            Dval = DctBlock[dblock, dblock]  # read dct value
            DMat[ih, iw] = Dval  # store it to 32x32 matrix

            iw = iw + 1
            # print('['+str(i)+','+str(j)+']')
        ih = ih + 1

    print('DCT Matrix of A band generated')

    ## Apply svd on Dmat and embed watermark image
    alpha = 0.1
    # [DMat2,U_SHL_w,V_SHL_w,Simg_temp]=WSVD(DMat,watermarkImage,alpha)
    Uimg, Simg, Vimg = np.linalg.svd(DMat, full_matrices=1, compute_uv=1)
    Simg = np.diag(Simg)  # convert to diagonal matrix
    Simg_temp = Simg  # save a copy of original Simg

    [x, y] = np.shape(watermarkImage)
    Simg = Simg + alpha * watermarkImage
    # SVD of watermarked S
    U_SHL_w, S_SHL_w, V_SHL_w = np.linalg.svd(Simg, full_matrices=1, compute_uv=1);
    # S_SHL_w=np.diag(S_SHL_w)

    ## Applying inverse SVD
    # DMat2 =Uimg* S_SHL_w * Vimg
    DMat2 = np.dot((Uimg * S_SHL_w), Vimg)

    print('SVD Applied on A band')
    A2 = A
    ih = 0;
    iw = 0;
    for i in range(0, Ha - 1, 8):
        iw = 0
        for j in range(0, Wa - 1, 8):
            Block = A2[i:i + 8, j:j + 8];
            DctBlock = cv2.dct(Block);
            DctBlock[dblock, dblock] = DMat2[ih, iw];
            Block2 = cv2.idct(DctBlock);
            A2[i:i + 8, j:j + 8] = Block2;
            iw = iw + 1;
        ih = ih + 1
    print('Inverse SVD Applied updated A band generated')

    # Perform inverse dwt and generate watermarked image
    coeffW = A2, (H, V, D)
    watermarkedImage = pywt.idwt2(coeffW, 'haar')
    print('Inverse DWT Applied')
    watermarkedImage = np.round_(watermarkedImage)
    watermarkedImage = np.uint8(watermarkedImage)

    # save watermarked image
    fn = 'DWTDCTSVD_Watermarked_' + FName  # set file name
    # cv2.imshow('DWTDCTSVD_Watermarked_' + FName, watermarkedImage)

    # CALCULATE MSE
    MSE = mse(coverImage, watermarkedImage)
    print('MSE (coverimage, watermarked)=' + str(MSE))
    PSNR = psnr(coverImage, watermarkedImage)
    print('PSNR (coverimage, watermarked)=' + str(PSNR))

    ## UPDATE EXCEL DATA
    Technique = 'DWT_DCT_SVD'
    XLS_path = "watermarking results.xlsx"
    XLS_Obj = openpyxl.load_workbook(XLS_path)
    XLS_Sheet_Obj = XLS_Obj.active
    CountROW = XLS_Sheet_Obj.max_row
    CountCol = 1
    print('ROW Location=' + str(CountROW))
    CELL1 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=1);
    CELL1.value = FName
    CELL2 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=2);
    CELL2.value = 'DWT_DCT_SVD'
    CELL3 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=3);
    CELL3.value = str(MSE)
    CELL4 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=4);
    CELL4.value = str(PSNR)
    XLS_Obj.save("watermarking results.xlsx")

    img = Image.fromarray(watermarkedImage)
    Floc = './result/DWTDCTSVD/' + fn  # set file path+file name
    img.save(Floc)  # save image at location
    print('Watermark Image displayed and saved at:' + Floc)


    return


# option 7
def ALL_RUN(coverImage, watermarkImage):
    print("All Test Run")
    print("1. -------------DWT WATERMARKING --------")
    DWT(coverImage, watermarkImage)
    print("\n\n2. -------------DCT WATERMARKING --------")
    I = coverImage;
    Imask = watermarkImage
    DCT(I, Imask)
    print("\n\n3. -------------DFT WATERMARKING --------")
    DFT(coverImage, watermarkImage)
    print("\n\n4. -------SVD WATERMARKING-----------")
    SVD(coverImage, watermarkImage)
    print("\n\n5. -------DWT_SVD WATERMARKING-----------")
    DWT_SVD(coverImage, watermarkImage)
    print("\n\n6. -------DWT_DCT_SVD WATERMARKING-----------")
    DWT_DCT_SVD(coverImage, watermarkImage)
#option 8
def ALL_RUN_except_DFT(coverImage, watermarkImage):
    print("All Test Run")
    print("1. -------------DWT WATERMARKING --------")
    DWT(coverImage, watermarkImage)
    print("\n\n2. -------------DCT WATERMARKING --------")
    I=coverImage;Imask=watermarkImage
    DCT(I, Imask)
    print("\n\n4. -------SVD WATERMARKING-----------")
    SVD(coverImage, watermarkImage)
    print("\n\n5. -------DWT_SVD WATERMARKING-----------")
    DWT_SVD(coverImage, watermarkImage)
    print("\n\n6. -------DWT_DCT_SVD WATERMARKING-----------")
    DWT_DCT_SVD(coverImage, watermarkImage)

if __name__ == "__main__":

    cv2.destroyAllWindows()  # close all previos window
    watermarkImage = cv2.imread('watermarkImage.JPG', 0)
    ##cv2.imshow('Original Watermark Image', watermarkImage)
    FuncName = ['DWT', 'DCT', 'DFT', 'SVD', 'DWT_SVD', 'DWT_DCT_SVD', 'DWT_DFT_SVD']
    options = {
        1: DWT,
        2: DCT,
        3: DFT,
        4: SVD,
        5: DWT_SVD,
        6: DWT_DCT_SVD,
        7: ALL_RUN,
        8: ALL_RUN_except_DFT
    }

    val = input('What type of embedding you want to perform? \
                \n1.DWT\
                \n2.DCT\
                \n3.DFT\
                \n4.SVD\
                \n5.DWT-SVD\
                \n6.SVD-DCT-DWT\
                \n7.ALL_RUN\
                \n8.ALL_RUN_except_DFT\
                \nEnter your option: ')

    watermarking_function = options.get(int(val), None)

    if watermarking_function:

        f = glob.glob('.\coverimages\*.jpg')  # search files with extension .jpg in same folder
        NFiles = len(f)  # find total files found
        print('Files In folder')
        print(f)  # print all file names

        for Fc in range(0, NFiles):  # run till all fies processed
            # Fc=0;
            st = f[Fc]  # select the file name
            s = st[14:len(st)]  # remove unnecessary heading of file extension
            FName = st[14:len(st)]

            coverImage = cv2.imread(st, 0)
            print('\n**----------------------------------------')
            print('Processing File (' + str(Fc + 1) + '/' + str(NFiles) + '): ' + st)  # print file name
            watermarking_function(coverImage, watermarkImage)
        exit(1)
    else:
        print("Invalid Option")
        exit(1)

